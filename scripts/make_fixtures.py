"""Generate test .xlsx fixtures for fastsheet IronCalc-fidelity validation.

Produces three files in ../fixtures/:
  basic.xlsx         — labels, numbers, simple formulas (SUM, IF, VLOOKUP)
  tables.xlsx        — an Excel Table with structured references
  array_formulas.xlsx — dynamic array formulas (SORT, FILTER, UNIQUE, SEQUENCE)

Run: python scripts/make_fixtures.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).resolve().parent.parent / "fixtures"
OUT.mkdir(exist_ok=True)


def make_basic():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Region"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Total"
    rows = [
        ("North", 100, 150),
        ("South", 200, 175),
        ("East",  120, 140),
        ("West",  90,  110),
    ]
    for i, (region, q1, q2) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=q1)
        ws.cell(row=i, column=3, value=q2)
        ws.cell(row=i, column=4, value=f"=B{i}+C{i}")
    ws["A6"] = "Total"
    ws["B6"] = "=SUM(B2:B5)"
    ws["C6"] = "=SUM(C2:C5)"
    ws["D6"] = "=SUM(D2:D5)"
    ws["A8"] = "Lookup"
    ws["B8"] = "South"
    ws["C8"] = '=VLOOKUP(B8,A2:D5,4,FALSE)'
    ws["A9"] = "Status"
    ws["B9"] = '=IF(D6>800,"good","low")'
    wb.save(OUT / "basic.xlsx")


def make_tables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["OrderID", "Customer", "Amount", "Tax"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    data = [
        (1001, "Acme",   500.00, "=C2*0.1"),
        (1002, "Globex", 320.50, "=C3*0.1"),
        (1003, "Initech", 1200.00, "=C4*0.1"),
        (1004, "Umbrella", 75.25, "=C5*0.1"),
    ]
    for i, row in enumerate(data, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=v)
    table = Table(displayName="Orders", ref="A1:D5")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True,
    )
    ws.add_table(table)
    ws["F1"] = "Sum of Amount"
    ws["G1"] = "=SUM(Orders[Amount])"
    ws["F2"] = "Sum of Tax"
    ws["G2"] = "=SUM(Orders[Tax])"
    wb.save(OUT / "tables.xlsx")


def make_array_formulas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dyn"
    ws["A1"] = "Names"
    names = ["Charlie", "Alice", "Bob", "Alice", "Diana", "Bob"]
    for i, n in enumerate(names, start=2):
        ws.cell(row=i, column=1, value=n)
    ws["C1"] = "Sorted"
    ws["C2"] = "=SORT(A2:A7)"
    ws["E1"] = "Unique"
    ws["E2"] = "=UNIQUE(A2:A7)"
    ws["G1"] = "Sequence"
    ws["G2"] = "=SEQUENCE(5)"
    ws["I1"] = "Filter A"
    ws["I2"] = '=FILTER(A2:A7,A2:A7="Alice")'
    wb.save(OUT / "array_formulas.xlsx")


if __name__ == "__main__":
    make_basic()
    make_tables()
    make_array_formulas()
    print("wrote fixtures to", OUT)
    for p in sorted(OUT.glob("*.xlsx")):
        print(" ", p.name, p.stat().st_size, "bytes")
